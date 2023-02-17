import logging
import re
from tempfile import NamedTemporaryFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

logger = logging.getLogger(__name__)


async def validate_string(string: str) -> bool:
    if (
            not string or
            not string.isdigit()
            # любые другие проверки
    ):
        return False
    return True


async def process_file_handler(file: list[bytes]) -> bytes:
    result_strings = []
    for string in file:
        string = string.decode('utf8')
        if not await validate_string(string):
            # если строка не прошла валидацию
            # что-нибудь с ней делаем
            string = ''
        result_strings.append(string)
    # записываем результат во временный файл и возвращаем
    with NamedTemporaryFile(mode='w+t') as file:
        file.writelines(result_strings)
        file.seek(0)
        return file.read()


async def process_excel_handler(excel: bytes) -> bytes:
    """Обработка или создание нового файла excel."""
    # обработать загруженную таблицу
    workbook = load_workbook(excel)
    worksheet = workbook.active
    values_list: list[tuple] = list(worksheet.values)
    for values in values_list:
        # обработка каждой строки таблицы
        pass
    # или создать новую таблицу
    # workbook = Workbook()
    # worksheet = workbook.active
    data = []  # любые данные, которые надо записать
    worksheet.title = 'Title'
    columns = ['column1', 'column2']
    row_num = 1
    for col_num, column in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column
        cell.alignment = Alignment(horizontal='center')
    max_value_length = 30
    for item in data:
        # заполнение таблицы данными
        # определение длины максимального значения
        value_length = len(item)
        if max_value_length < value_length:
            max_value_length = value_length
    worksheet.column_dimensions['A'].width = max_value_length
    worksheet.column_dimensions['B'].width = max_value_length
    # вывод excel таблицы с помощью временного файла
    with NamedTemporaryFile() as file:
        workbook.save(file.name)
        file.seek(0)
        return file.read()
